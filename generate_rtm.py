import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# RTM Data
data = [
    ["FRQ1","Ingest customer interaction data","Data ingestion module","Test ingestion from multiple sources","Covered"],
    ["FRQ2","Clean text, detect language, handle informal/coded","Preprocessing module","Test preprocessing pipeline","Covered"],
    ["FRQ3","Classify text into sentiment","Sentiment analysis module","Test sentiment classification","Covered"],
    ["FRQ4","Categorize feedback into topics","Aspect/topic classification module","Test topic labeling","Covered"],
    ["FRQ5","Real-time dashboard updates","Streamlit dashboard","Submit messages; verify dashboard","Covered"],
    ["FRQ6","Near real-time analysis","Live feedback engine","Test session/live updates","Covered"],
    ["FRQ7","Attach sentiment/topic labels","Data processing module","Verify all messages labeled","Covered"],
    ["FRQ8","Generate analytical reports","Reporting module","Generate PDF/CSV; check summaries","Covered"],
    ["FRQ9","Store raw feedback and results","Database module","Insert & retrieve sample data","Covered"],
    ["NFRQ1","Achieve ≥90% accuracy","Model evaluation","Test model on labeled dataset","Pending"],
    ["NFRQ2","Consistent performance across sources","Model evaluation","Compare accuracy across sources","Pending"],
    ["NFRQ3","Explain sentiment predictions","Model explainability","Verify confidence scores & explanations","Covered"],
    ["NFRQ4","Handle increasing data volume","Scalability tests","Test large datasets; check performance","Pending"],
    ["NFRQ5","Use anonymized data","Preprocessing & storage","Verify PII removed","Covered"],
    ["NFRQ6","Do not store PII","Storage compliance","Verify no personal identifiers stored","Covered"],
    ["NFRQ7","Dashboard understandable by staff","Streamlit dashboard","UX review by sample users","Pending"],
    ["NFRQ8","Bias awareness","Model evaluation","Check predictions across demographics","Pending"],
    ["NFRQ9","Transparent model behavior","Logging / explainability","Verify logs & confidence scores","Covered"],
    ["NFRQ10","Data protection compliance","DB & security","Check anonymization & secure access","Covered"],
]

# Create DataFrame
df = pd.DataFrame(data, columns=["Req ID","Requirement Description","Functional Component","Test Case / Verification","Status / Notes"])

# Save to Excel
filename = "Sentiment_RTM.xlsx"
df.to_excel(filename, index=False)

# Load workbook for formatting
wb = load_workbook(filename)
ws = wb.active

# Define colors
color_map = {
    "Covered": "C6EFCE",  # Green
    "Pending": "FFEB9C",  # Yellow
    "Failed": "FFC7CE"    # Red
}

# Apply fill colors
for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=5).value
    fill_color = color_map.get(status, None)
    if fill_color:
        ws.cell(row=row, column=5).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Save workbook
wb.save(filename)
print(f"Color-coded RTM saved as '{filename}'")